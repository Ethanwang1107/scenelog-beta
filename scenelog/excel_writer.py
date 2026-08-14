"""Excel 场记表输出 — pandas + openpyxl"""

from __future__ import annotations

import logging
import os
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from scenelog.config import EXCEL_FILE, TRANSCRIPTS_DIR

logger = logging.getLogger(__name__)

# 场记表列定义
COLUMNS = [
    "文件名",
    "相对路径",
    "时长",
    "拍摄时间",
    "时间码",
    "分辨率",
    "帧率",
    "编码",
    "内容摘要",
    "关键词",
    "逐字稿",
    "处理状态",
    "人物",
    "地点",
    "可用点",
]

# 自动填充列
AUTO_COLUMNS = COLUMNS[:12]  # 前 12 列自动生成
MANUAL_COLUMNS = COLUMNS[12:]  # 后 3 列手填

# 逐字稿列索引（0-based）
TRANSCRIPT_COL_IDX = COLUMNS.index("逐字稿")  # = 10


def load_excel_edits(excel_path: Path) -> dict[str, dict]:
    """按相对路径读取用户在现有场记表中修改的可编辑字段。"""
    if not excel_path.exists():
        return {}
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=False)
        worksheet = workbook["场记表"]
        headers = {
            str(cell.value): index
            for index, cell in enumerate(worksheet[1], start=1)
            if cell.value
        }
        required = {"相对路径", "内容摘要", "关键词", *MANUAL_COLUMNS}
        if not required.issubset(headers):
            return {}

        edits = {}
        for row_index in range(2, worksheet.max_row + 1):
            rel_path = _cell_text(
                worksheet.cell(row=row_index, column=headers["相对路径"]).value
            )
            if not rel_path:
                continue
            keyword_text = _cell_text(
                worksheet.cell(row=row_index, column=headers["关键词"]).value
            )
            edits[rel_path] = {
                "内容摘要": _cell_text(
                    worksheet.cell(row=row_index, column=headers["内容摘要"]).value
                ),
                "关键词": _split_keywords(keyword_text),
                **{
                    column: _cell_text(
                        worksheet.cell(row=row_index, column=headers[column]).value
                    )
                    for column in MANUAL_COLUMNS
                },
            }
        return edits
    except Exception as e:
        logger.warning("读取现有场记表修正失败，将使用状态文件结果: %s", e)
        return {}


def generate_excel(
    records: list[dict],
    output_dir: Path,
    transcripts_dir: Path,
    scenelog_dir: Path | None = None,
) -> Path:
    """生成场记表 Excel 文件。"""
    if not records:
        logger.warning("无素材记录，跳过 Excel 生成")
        return output_dir / EXCEL_FILE

    excel_path = output_dir / EXCEL_FILE
    excel_edits = load_excel_edits(excel_path)
    rows = []
    transcript_paths: list[Path | None] = []  # 记录每行对应的 SRT 绝对路径

    for rec in records:
        meta = rec.get("metadata", {})
        summary = rec.get("summary", {})
        state = rec.get("state", {})

        material_id = rec.get("material_id", "")
        file_name = meta.get("file_name", rec.get("file_name", ""))
        rel_path = meta.get("rel_path", rec.get("rel_path", ""))

        speaker_transcript = state.get("speaker_transcript_srt", "")
        srt_abs = (
            scenelog_dir / speaker_transcript
            if scenelog_dir and speaker_transcript
            else transcripts_dir / f"{material_id}.srt"
        )
        try:
            srt_rel = str(srt_abs.relative_to(output_dir))
        except ValueError:
            srt_rel = str(srt_abs)
        existing_manual = excel_edits.get(rel_path, {})
        automatic_people = ", ".join(rec.get("people", []))
        existing_people = existing_manual.get("人物", "")
        previous_automatic_people = state.get("auto_people_text", "")
        people_text = (
            existing_people
            if (
                rel_path in excel_edits
                and existing_people != previous_automatic_people
            )
            else automatic_people
        )

        # 读取转录文本摘要（不读全文）
        transcript_display = ""
        if state.get("transcription") == "skipped":
            transcript_display = "[未识别出语音内容]"
            transcript_paths.append(None)
        elif srt_abs.exists():
            try:
                text = srt_abs.read_text(encoding="utf-8").strip()
                if text and "[未识别出语音内容]" not in text:
                    transcript_display = srt_rel
                    transcript_paths.append(srt_abs)
                else:
                    transcript_display = "[未识别出语音内容]"
                    transcript_paths.append(None)
            except Exception:
                transcript_display = "[读取失败]"
                transcript_paths.append(None)
        else:
            transcript_display = "[未识别出语音内容]"
            transcript_paths.append(None)

        row = {
            "文件名": meta.get("file_name", rec.get("file_name", "")),
            "相对路径": rel_path,
            "时长": meta.get("duration_str", ""),
            "拍摄时间": meta.get("creation_time", ""),
            "时间码": meta.get("timecode", ""),
            "分辨率": _format_resolution(meta),
            "帧率": meta.get("frame_rate", ""),
            "编码": _format_codec(meta),
            "内容摘要": summary.get("摘要", ""),
            "关键词": ", ".join(summary.get("关键词", [])) if summary.get("关键词") else "",
            "逐字稿": transcript_display,
            "处理状态": _format_status(state),
            "人物": people_text,
            "地点": existing_manual.get("地点", ""),
            "可用点": existing_manual.get("可用点", ""),
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMNS)

    # 原子写入
    tmp_dir = scenelog_dir or output_dir
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_fd, tmp_path = tempfile.mkstemp(
        dir=str(tmp_dir), prefix=".excel_", suffix=".xlsx"
    )
    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="场记表", index=False)

            ws = writer.sheets["场记表"]

            # 为逐字稿列添加超链接（覆盖 pandas 写入的纯文本）
            transcript_col_letter = get_column_letter(TRANSCRIPT_COL_IDX + 1)
            link_font = Font(color="0563C1", underline="single")

            for row_idx, srt_path in enumerate(transcript_paths, start=2):  # 第 2 行开始
                cell = ws.cell(row=row_idx, column=TRANSCRIPT_COL_IDX + 1)
                if srt_path:
                    cell.hyperlink = srt_path.as_uri()
                    cell.font = link_font
                    cell.value = cell.value or "[点击打开逐字稿]"

            # 列宽
            col_widths = {
                "A": 22, "B": 30, "C": 10, "D": 20, "E": 12,
                "F": 12, "G": 10, "H": 14, "I": 40, "J": 24,
                "K": 50, "L": 14, "M": 12, "N": 12, "O": 16,
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 冻结首行
            ws.freeze_panes = "A2"

            # 自动筛选
            ws.auto_filter.ref = f"A1:O{len(rows) + 1}"

        os.replace(tmp_path, excel_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise

    logger.info("场记表已生成: %s (%d 条记录)", excel_path, len(rows))
    return excel_path


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_keywords(value: str) -> list[str]:
    normalized = value.replace("，", ",").replace("、", ",")
    return [keyword.strip() for keyword in normalized.split(",") if keyword.strip()]


def _format_resolution(meta: dict) -> str:
    w = meta.get("width", 0)
    h = meta.get("height", 0)
    if w and h:
        return f"{w}x{h}"
    return ""


def _format_codec(meta: dict) -> str:
    parts = []
    if meta.get("video_codec"):
        parts.append(meta["video_codec"])
    if meta.get("audio_codec"):
        parts.append(meta["audio_codec"])
    return "/".join(parts) if parts else ""


def _format_status(state: dict) -> str:
    """格式化处理状态为简洁可读字符串。"""
    errors = state.get("errors", {})
    for step in ["metadata", "audio_extract"]:
        if state.get(step) == "failed_terminal":
            return f"失败 ({errors.get(step, state.get('last_error', step))})"

    if state.get("transcription") == "failed_retryable":
        return f"失败 ({errors.get('transcription', '转录失败')})"

    if state.get("summary") == "failed_retryable":
        return f"失败 ({errors.get('summary', '摘要失败')})"

    if state.get("vision") == "failed_retryable":
        return f"失败 ({errors.get('vision', '画面理解失败')})"

    if state.get("people") == "failed_retryable":
        return f"失败 ({errors.get('people', '人物识别失败')})"

    if state.get("index") == "failed_retryable":
        return f"失败 ({errors.get('index', '索引失败')})"

    summary_text = state.get("summary_text", "")
    if summary_text == "[无语音]":
        return "无语音"

    if state.get("audio_extract") == "skipped" and not state.get("vision_summary"):
        return "无音轨"

    core_steps = [
        "metadata",
        "audio_extract",
        "transcription",
        "vision",
        "people",
        "summary",
        "index",
    ]
    all_ok = all(
        state.get(s) in ("success", "skipped")
        for s in core_steps
    )
    if all_ok:
        return "成功"

    if state.get("vad") == "skipped" and not summary_text:
        return "无语音"

    return "处理中"

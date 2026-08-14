from pathlib import Path

import openpyxl

from scenelog.excel_writer import generate_excel, load_excel_edits
from scenelog.scanner import scan_media


def _record(material_id: str = "mat_001") -> dict:
    return {
        "material_id": material_id,
        "metadata": {
            "file_name": "C001.MOV",
            "rel_path": "card-a/C001.MOV",
            "duration_str": "00:00:10",
        },
        "summary": {"摘要": "院子里一名男子劈柴", "关键词": ["院子", "劈柴"]},
        "state": {
            "metadata": "success",
            "audio_extract": "success",
            "vad": "success",
            "transcription": "success",
            "vision": "success",
            "summary": "success",
            "index": "success",
            "excel": "pending",
        },
    }


def test_excel_preserves_manual_columns_and_uses_material_id_transcript(tmp_path: Path):
    transcripts = tmp_path / "transcripts"
    transcripts.mkdir()
    (transcripts / "mat_001.srt").write_text(
        "1\n00:00:00,000 --> 00:00:01,000\n测试对白\n\n",
        encoding="utf-8",
    )
    excel_path = generate_excel([_record()], tmp_path, transcripts, tmp_path)

    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook["场记表"]
    worksheet["M2"] = "张三"
    worksheet["N2"] = "村口"
    worksheet["O2"] = "00:03可用"
    workbook.save(excel_path)

    generate_excel([_record()], tmp_path, transcripts, tmp_path)
    worksheet = openpyxl.load_workbook(excel_path)["场记表"]

    assert [worksheet["M2"].value, worksheet["N2"].value, worksheet["O2"].value] == [
        "张三",
        "村口",
        "00:03可用",
    ]
    assert worksheet["K2"].value == "transcripts/mat_001.srt"
    assert worksheet["K2"].hyperlink.target.endswith("/transcripts/mat_001.srt")


def test_scanner_excludes_custom_output_directory(tmp_path: Path):
    source = tmp_path / "source"
    custom_output = source / "results"
    source.mkdir()
    custom_output.mkdir()
    (source / "clip.mov").write_bytes(b"source")
    (custom_output / "derived.wav").write_bytes(b"derived")

    found = scan_media(source, excluded_dirs=[custom_output])

    assert [path.relative_to(source).as_posix() for path in found] == ["clip.mov"]


def test_excel_edits_include_summary_keywords_and_manual_columns(tmp_path: Path):
    excel_path = generate_excel([_record()], tmp_path, tmp_path, tmp_path)
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook["场记表"]
    worksheet["I2"] = "人工改写的摘要"
    worksheet["J2"] = "院子，人物、劈柴"
    worksheet["M2"] = "张三"
    workbook.save(excel_path)

    edits = load_excel_edits(excel_path)

    assert edits["card-a/C001.MOV"] == {
        "内容摘要": "人工改写的摘要",
        "关键词": ["院子", "人物", "劈柴"],
        "人物": "张三",
        "地点": "",
        "可用点": "",
    }


def test_excel_edit_reader_does_not_truncate_generated_identity_keywords(
    tmp_path: Path,
):
    record = _record()
    record["summary"]["关键词"] = [
        "王书记",
        "老刘",
        "乡村小路",
        "行走",
        "交谈",
        "冲锋衣",
        "蓝天",
        "山脉",
    ]
    excel_path = generate_excel([record], tmp_path, tmp_path, tmp_path)

    edits = load_excel_edits(excel_path)

    assert edits["card-a/C001.MOV"]["关键词"] == record["summary"]["关键词"]


def test_excel_updates_automatic_people_but_preserves_manual_people(tmp_path: Path):
    record = _record()
    record["people"] = ["人物01"]
    record["state"]["auto_people_text"] = ""
    excel_path = generate_excel([record], tmp_path, tmp_path, tmp_path)

    record["people"] = ["张三"]
    record["state"]["auto_people_text"] = "人物01"
    generate_excel([record], tmp_path, tmp_path, tmp_path)
    worksheet = openpyxl.load_workbook(excel_path)["场记表"]
    assert worksheet["M2"].value == "张三"

    worksheet["M2"] = "导演手工标注"
    worksheet.parent.save(excel_path)
    record["people"] = ["李四"]
    record["state"]["auto_people_text"] = "张三"
    generate_excel([record], tmp_path, tmp_path, tmp_path)
    worksheet = openpyxl.load_workbook(excel_path)["场记表"]
    assert worksheet["M2"].value == "导演手工标注"

    worksheet["M2"] = ""
    worksheet.parent.save(excel_path)
    record["state"]["auto_people_text"] = "李四"
    generate_excel([record], tmp_path, tmp_path, tmp_path)
    worksheet = openpyxl.load_workbook(excel_path)["场记表"]
    assert worksheet["M2"].value is None
